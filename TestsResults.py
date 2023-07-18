""" I do not know if we need this, Rojeh, I am waiting for your opinion."""

import openpyxl


def get_row(sheet, row_name, session_num):
    column_name = 'Number'

    column_index = None
    for column in sheet.iter_cols():
        if column[0].value == column_name:
            if row_name is None:
                return column[0].column
            column_index = column[0].column
            break

    row_index = 1
    for patient in sheet.iter_rows(min_col=column_index, max_col=column_index):
        if patient[0].value == row_name:
            if session_num is None:
                return row_index
            else:
                cell = sheet.cell(row=row_index, column=column_index + 1)
                if cell.value == session_num:
                    return row_index
        row_index += 1

    return -1

def get_patient_details(vid):
    workbook = openpyxl.load_workbook('Data.xlsx')
    for row in workbook['Sheet1']:
        if row[1].value == vid['name'] and row[2].value:
            return row

def get_start_time(row_name, session_num):
    workbook = openpyxl.load_workbook('Data.xlsx')
    sheet = workbook['Sheet1']

    row_index = get_row(sheet, row_name, session_num)

    column_name = 'TimeToCut'
    column_index = None
    for column in sheet.iter_cols():
        if column[0].value == column_name:
            column_index = column[0].column
            break

    cell = sheet.cell(row=row_index, column=column_index)
    cell_val = cell.value
    cell_val = cell_val[:len(cell_val)-1]
    nums = cell_val.split(":")
    minutes = int(nums[0])
    seconds = int(nums[1])

    return (60 * minutes) + seconds


def get_real_speed(row_name, session_num):
    workbook = openpyxl.load_workbook('SPSS.xlsx')
    sheet = workbook['RESULTS']

    row_index = get_row(sheet, row_name, None)

    column_name = '4MWTspeed_' + session_num
    column_index = None
    for column in sheet.iter_cols():
        if column[0].value == column_name:
            column_index = column[0].column
            break

    cell = sheet.cell(row=row_index, column=column_index)
    cell_val = cell.value

    return cell_val

def get_real_computed_speed(row_name, session_num):
    workbook = openpyxl.load_workbook('Data.xlsx')
    sheet = workbook.active

    row_index = get_row(sheet, row_name, None) + session_num
    print(row_index)

    column_name = 'Computed Speed'
    column_index = None
    for column in sheet.iter_cols():
        if column[0].value == column_name:
            column_index = column[0].column
            break

    cell = sheet.cell(row=row_index, column=column_index)
    cell_val = cell.value

    return cell_val


def save_evaluation(path, value_to_save, column_name='Computed Speed', kerem=False):
    workbook = openpyxl.load_workbook('Data.xlsx')

    # Select the worksheet
    worksheet = workbook.active

    for i in range(1, worksheet.max_column + 1):
        if worksheet.cell(row=1, column=i).value == column_name:
            column_index = i
            break

    # Get the existing column
    if kerem:
        for i in range(1, worksheet.max_row):
            if worksheet.cell(row=i, column=2).value == path:
                cell_value = worksheet.cell(row=i, column=column_index).value
                print(cell_value)
                if column_name == 'Computed Speed' and (cell_value is not None or cell_value != ""):
                    worksheet.cell(row=i, column=column_index, value=value_to_save)
                elif (column_name == 'Start Line' or column_name == 'End Line') and value_to_save is not None:
                    data_string = ', '.join(str(item) for item in value_to_save)
                    worksheet.cell(row=i, column=column_index, value=data_string)
                elif column_name == 'Start Line' or column_name == 'End Line':
                    if cell_value is None or cell_value == "":
                        return None
                    data = [int(item) for item in cell_value.split(', ')]
                    return data
    else:
        nameOf = path.split('/')
        nameOf = nameOf[-1].split('_')
        nameOf[1] = nameOf[1].split('.')[0]
        print(nameOf)
        for i in range(1, worksheet.max_row):
            if worksheet.cell(row=i, column=1).value == int(nameOf[0]) and \
                    worksheet.cell(row=i, column=2).value == nameOf[1]:
                cell_value = worksheet.cell(row=i, column=column_index).value
                print(cell_value)
                if column_name == 'Computed Speed' and (cell_value is not None or cell_value != ""):
                    worksheet.cell(row=i, column=column_index, value=value_to_save)
                elif (column_name == 'Start Line' or column_name == 'End Line') and value_to_save is not None:
                    data_string = ', '.join(str(item) for item in value_to_save)
                    worksheet.cell(row=i, column=column_index, value=data_string)
                elif column_name == 'Start Line' or column_name == 'End Line':
                    if cell_value is None or cell_value == "":
                        return None
                    data = [int(item) for item in cell_value.split(', ')]
                    return data

    # Save the modified Excel file
    workbook.save('Data.xlsx')


